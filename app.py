"""
Aircon Reconciliation Web App
Flask backend that accepts file uploads, runs the original aircon_check_final.py
reconciliation, and serves results.
"""

import os
import re
import sys
import uuid
import shutil
import math
from pathlib import Path

import io
from collections import defaultdict
from calendar import month_name

from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import load_workbook

# Lazy import of reconcile() — deferred to first use so app can start even if deps have issues
_reconcile = None

def get_reconcile():
    global _reconcile
    if _reconcile is None:
        try:
            from aircon_check_final import reconcile
            _reconcile = reconcile
        except ImportError:
            PARENT_DIR = str(Path(__file__).resolve().parent.parent)
            if PARENT_DIR not in sys.path:
                sys.path.insert(0, PARENT_DIR)
            from aircon_check_final import reconcile
            _reconcile = reconcile
    return _reconcile

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
)

# Store temp session data: session_id -> {folder_path, output_path, output_filename}
SESSIONS = {}

# Base temp directory — use system temp on Railway, local subfolder otherwise
import tempfile
TEMP_BASE = Path(tempfile.gettempdir()) / "aircon_reconciliation"
TEMP_BASE.mkdir(exist_ok=True)


def parse_summary_text(summary_text):
    """Parse the text summary returned by reconcile() into a structured dict."""
    def extract_int(pattern, text, default=0):
        m = re.search(pattern, text)
        return int(m.group(1)) if m else default

    def extract_float(pattern, text, default=0.0):
        m = re.search(pattern, text)
        return float(m.group(1)) if m else default

    return {
        "matched": extract_int(r"Matched cleanly:\s*(\d+)", summary_text),
        "zero_charge": extract_int(r"Zero-charge \(review\):\s*(\d+)", summary_text),
        "unclear": extract_int(r"Unclear \(review\):\s*(\d+)", summary_text),
        "mismatch": extract_int(r"Amount mismatch:\s*(\d+)", summary_text),
        "missing": extract_int(r"Missing from PDFs:\s*(\d+)", summary_text),
        "total_landlord_rows": extract_int(r"Landlord rows:\s*(\d+)", summary_text),
        "unbilled_pdfs": extract_int(r"PDF bookings not billed:\s*(\d+)", summary_text),
    }


def to_float(value, default=0.0):
    """Parse Excel numeric/currency values safely."""
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("$", "").replace(",", "").strip()
    if not text or text in {"-", "–", "—"}:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def normalise_status(value):
    return str(value or "").strip().upper()


def find_header_row(ws, required_headers, max_scan_rows=8):
    """Find the row containing the expected headers, allowing intro/key rows."""
    required = {header.lower() for header in required_headers}
    for row_num in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = {
            str(cell.value or "").strip().lower()
            for cell in ws[row_num]
            if cell.value is not None
        }
        if required.issubset(values):
            return row_num
    return 1


def get_row_dicts(ws, header_row=1, required_headers=None):
    """Return worksheet rows as dictionaries keyed by header text."""
    if required_headers:
        header_row = find_header_row(ws, required_headers)

    headers = [
        "" if cell.value is None else str(cell.value).strip()
        for cell in ws[header_row]
    ]
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return rows


def infer_month_key(filename, recon_rows):
    """Infer YYYY-MM and display label from workbook rows or filename."""
    for row in recon_rows:
        value = row.get("Date")
        if hasattr(value, "year") and hasattr(value, "month"):
            return f"{value.year}-{value.month:02d}", f"{month_name[value.month]} {value.year}"
        text = str(value or "")
        m = re.search(
            r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})",
            text,
            flags=re.I,
        )
        if m:
            month_word = m.group(2).lower()[:3]
            month_lookup = {
                "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
            }
            if month_word in month_lookup:
                month_num = month_lookup[month_word]
                year = int(m.group(3))
                return f"{year}-{month_num:02d}", f"{month_name[month_num]} {year}"

    m = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[ _-]*(\d{4})",
        filename,
        flags=re.I,
    )
    if m:
        word = m.group(1).lower()[:3]
        month_lookup = {
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
            "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
        }
        month_num = month_lookup.get(word)
        if month_num:
            year = int(m.group(2))
            return f"{year}-{month_num:02d}", f"{month_name[month_num]} {year}"

    return filename, filename


# ── Room grouping ──────────────────────────────────────────────────────────────
# Maps individual room numbers to a canonical display label.
# Rooms on the same floor that are always booked together share a group.
ROOM_GROUPS = {
    700:  "Room 700, Auditorium (L7 & L8)",
    701:  "Rooms 701/702",
    702:  "Rooms 701/702",
    801:  "Rooms 801/802",
    802:  "Rooms 801/802",
    901:  "Rooms 901/902",
    902:  "Rooms 901/902",
    903:  "Rooms 901/902",
    10:   "Rooms 1001/1002",
    1001: "Rooms 1001/1002",
    1002: "Rooms 1001/1002",
    1003: "Rooms 1001/1002",
    1104: "Rooms 1104/1105",
    1105: "Rooms 1104/1105",
}


def clean_room_label(value):
    """Map a raw PDF location string to a canonical room group.

    Strategy:
      1. Extract all numbers from the text.
      2. Look up 3-4 digit room numbers in ROOM_GROUPS first (most specific).
      3. Fall back to smaller numbers (e.g. '10' for Level 10).
      4. Detect 'mezzanine' or '7M' keywords (after room-number check).
      5. If nothing matches, return the cleaned text as-is.
    """
    text = str(value or "").strip()
    if not text:
        return "Unknown"

    # Extract every number from the text
    all_numbers = [int(n) for n in re.findall(r"\d+", text)]

    # Try 3-4 digit room numbers first (most reliable identifiers)
    for num in all_numbers:
        if 100 <= num <= 9999 and num in ROOM_GROUPS:
            return ROOM_GROUPS[num]

    # Try smaller numbers (e.g. standalone "10" for Level 10)
    for num in all_numbers:
        if num in ROOM_GROUPS:
            return ROOM_GROUPS[num]

    # Keyword match: mezzanine / 7M (checked after room numbers)
    if re.search(r"mezzanine|(?:^|\b)7\s*M(?:\b|$)", text, re.I):
        return "Mezzanine"

    # Fallback: join any 3-4 digit numbers found
    room_numbers = sorted(set(n for n in all_numbers if 100 <= n <= 9999))
    if room_numbers:
        return "/".join(str(n) for n in room_numbers)

    return re.sub(r"\s+", " ", text)


def clean_room_labels(value):
    """Return separate room labels for room cost ranking from PDF rows."""
    text = str(value or "").strip()
    if not text:
        return ["Unknown"]

    normalized = re.sub(r"\s+", " ", text)
    compact = re.sub(r"[^a-z0-9]", "", normalized.lower())

    if re.search(r"mezzanine|(?:^|\b)7\s*M(?:\b|$)", normalized, re.I) or "level7m" in compact:
        return ["Mezzanine Level 7M"]

    room_aliases = {
        700: ["700", "70 0", "7 00", "auditorium"],
        701: ["701", "70 1", "7 01"],
        702: ["702", "70 2", "7 02"],
        801: ["801", "80 1", "8 01", "802", "80 2", "8 02"],
        901: ["901", "90 1", "9 01"],
        902: ["902", "90 2", "9 02"],
        903: ["903", "90 3", "9 03"],
        1001: ["1001", "10 01"],
        1002: ["1002", "10 02"],
        1003: ["1003", "10 03"],
        1104: ["1104", "11 04", "room 104"],
        1105: ["1105", "11 05", "room 105"],
    }
    room_labels = {
        700: "Room 700, Auditorium (L7 & L8)",
        701: "Room 701",
        702: "Room 702",
        801: "Room 801",
        901: "Room 901",
        902: "Room 902",
        903: "Room 903",
        1001: "Room 1001",
        1002: "Room 1002",
        1003: "Room 1003",
        1104: "Room 1104",
        1105: "Room 1105",
    }

    matched_rooms = []
    for room, aliases in room_aliases.items():
        for alias in aliases:
            alias_compact = re.sub(r"[^a-z0-9]", "", alias.lower())
            if alias_compact and alias_compact in compact:
                matched_rooms.append(room)
                break

    if "level9" in compact and not any(room in matched_rooms for room in (901, 902, 903)):
        matched_rooms.extend([901, 902, 903])

    if "level10" in compact:
        for room in (1001, 1002):
            if room not in matched_rooms:
                matched_rooms.append(room)
        if "1003" in compact and 1003 not in matched_rooms:
            matched_rooms.append(1003)

    if "level13" in compact or any(token in compact for token in ("1304", "1305", "13041305")):
        return ["Level 13"]

    if matched_rooms:
        return [room_labels[room] for room in sorted(set(matched_rooms), key=matched_rooms.index)]

    if "level11" in compact:
        return ["Room 1104", "Room 1105"]

    return [normalized]



def forecast_billing(months_data, forecast_horizon=None):
    """
    Predict future monthly billing from uploaded reconciled reports.

    Main method (Year-on-year run-rate):
      - Use the latest uploaded year as the current year.
      - Compare current-year months against the same months from the previous year.
      - Apply that year-on-year growth factor to future months from the previous year.

    Fallback (Seasonal trend):
      - If there is not enough same-month prior-year data, use linear regression
        with per-month seasonal adjustment so the forecast follows the spending
        curve rather than a flat line.

    forecast_horizon: If None, automatically extends through December of the
    current data year to ensure a full year-on-year forecast.
    """
    if len(months_data) < 3:
        return {"items": [], "total": 0.0, "method": "Need at least 3 months of data"}

    parsed = []
    for item in months_data:
        try:
            year, month = [int(part) for part in item["key"].split("-")]
        except Exception:
            continue
        parsed.append({
            "year": year,
            "month": month,
            "key": item["key"],
            "label": item["label"],
            "billed": float(item.get("billed") or 0),
        })

    if len(parsed) < 3:
        return {"items": [], "total": 0.0, "method": "Need month-labelled data"}

    parsed.sort(key=lambda item: (item["year"], item["month"]))
    by_month = {(item["year"], item["month"]): item["billed"] for item in parsed}

    current_year = max(item["year"] for item in parsed)
    current_months = [item["month"] for item in parsed if item["year"] == current_year]
    if not current_months:
        return {"items": [], "total": 0.0, "method": "Need current-year data"}

    latest_month = max(current_months)

    # Auto-compute horizon to always reach December of the current year
    if forecast_horizon is None:
        forecast_horizon = max(6, 12 - latest_month)
    previous_year = current_year - 1
    comparable_months = [
        month for month in current_months
        if (previous_year, month) in by_month and by_month[(previous_year, month)] > 0
    ]

    forecasts = []
    residuals = []
    method = "Year-on-year run-rate"
    growth_factor = None

    if comparable_months:
        current_total = sum(by_month[(current_year, month)] for month in comparable_months)
        previous_total = sum(by_month[(previous_year, month)] for month in comparable_months)
        growth_factor = current_total / previous_total if previous_total else 1.0

        for month in comparable_months:
            expected = by_month[(previous_year, month)] * growth_factor
            residuals.append(by_month[(current_year, month)] - expected)

        for step in range(1, forecast_horizon + 1):
            future_month = latest_month + step
            future_year = current_year + (future_month - 1) // 12
            future_month = ((future_month - 1) % 12) + 1

            prior_value = by_month.get((future_year - 1, future_month))
            if prior_value is None:
                prior_value = by_month.get((previous_year, future_month))

            if prior_value is None:
                break

            predicted = max(0, prior_value * growth_factor)
            key = f"{future_year}-{future_month:02d}"
            forecasts.append({
                "key": key,
                "label": f"{month_name[future_month]} {future_year}",
                "billed": round(predicted, 2),
                "source": f"{future_year - 1} same month x {growth_factor:.2f}",
            })
    else:
        method = "Seasonal trend"
        values = [item["billed"] for item in parsed]
        n = len(values)
        xs = list(range(n))

        sum_x = sum(xs)
        sum_y = sum(values)
        sum_xy = sum(x * y for x, y in zip(xs, values))
        sum_x2 = sum(x * x for x in xs)

        denom = n * sum_x2 - sum_x * sum_x
        if denom == 0:
            slope = 0.0
            intercept = sum_y / n
        else:
            slope = (n * sum_xy - sum_x * sum_y) / denom
            intercept = (sum_y - slope * sum_x) / n

        # Compute seasonal factors (average residual per month-of-year)
        seasonal_sum = defaultdict(float)
        seasonal_count = defaultdict(int)
        for i, item in enumerate(parsed):
            trend_val = slope * i + intercept
            seasonal_sum[item["month"]] += item["billed"] - trend_val
            seasonal_count[item["month"]] += 1

        seasonal_factor = {}
        for moy in range(1, 13):
            seasonal_factor[moy] = (
                seasonal_sum[moy] / seasonal_count[moy]
                if seasonal_count[moy] > 0
                else 0.0
            )

        # Residuals with seasonal adjustment (for confidence bands)
        for i, item in enumerate(parsed):
            fitted = slope * i + intercept + seasonal_factor.get(item["month"], 0.0)
            residuals.append(item["billed"] - fitted)

        last = parsed[-1]
        for step in range(1, forecast_horizon + 1):
            future_month = last["month"] + step
            future_year = last["year"] + (future_month - 1) // 12
            future_month = ((future_month - 1) % 12) + 1
            trend_prediction = slope * (n - 1 + step) + intercept
            seasonal_adj = seasonal_factor.get(future_month, 0.0)
            predicted = max(0, trend_prediction + seasonal_adj)
            key = f"{future_year}-{future_month:02d}"
            forecasts.append({
                "key": key,
                "label": f"{month_name[future_month]} {future_year}",
                "billed": round(predicted, 2),
                "source": "Seasonal projection",
            })

    if len(residuals) > 1:
        mean_r = sum(residuals) / len(residuals)
        std_r = math.sqrt(sum((r - mean_r) ** 2 for r in residuals) / (len(residuals) - 1))
    else:
        std_r = 0.0

    for item in forecasts:
        band = max(std_r * 1.5, item["billed"] * 0.12)
        item["lower"] = round(max(0, item["billed"] - band), 2)
        item["upper"] = round(item["billed"] + band, 2)

    return {
        "items": forecasts,
        "total": round(sum(item["billed"] for item in forecasts), 2),
        "method": method,
        "growth_factor": round(growth_factor, 4) if growth_factor is not None else None,
        "history_months": len(parsed),
    }


def legacy_forecast_billing(months_data, forecast_horizon=6):
    """Older trend-only forecast retained for reference."""
    if len(months_data) < 3:
        return []

    values = []
    month_of_year = []
    for i, m in enumerate(months_data):
        values.append(m["billed"])
        parts = m["key"].split("-")
        month_of_year.append(int(parts[1]))

    n = len(values)
    xs = list(range(n))

    # --- Linear regression: y = slope * x + intercept ---
    sum_x = sum(xs)
    sum_y = sum(values)
    sum_xy = sum(x * y for x, y in zip(xs, values))
    sum_x2 = sum(x * x for x in xs)

    denom = n * sum_x2 - sum_x * sum_x
    if denom == 0:
        slope = 0.0
        intercept = sum_y / n
    else:
        slope = (n * sum_xy - sum_x * sum_y) / denom
        intercept = (sum_y - slope * sum_x) / n

    # --- Seasonal factors (average residual per month-of-year) ---
    seasonal_sum = defaultdict(float)
    seasonal_count = defaultdict(int)
    for i, val in enumerate(values):
        trend_val = slope * i + intercept
        residual = val - trend_val
        seasonal_sum[month_of_year[i]] += residual
        seasonal_count[month_of_year[i]] += 1

    seasonal_factor = {}
    for moy in range(1, 13):
        if seasonal_count[moy] > 0:
            seasonal_factor[moy] = seasonal_sum[moy] / seasonal_count[moy]
        else:
            seasonal_factor[moy] = 0.0

    # --- Generate forecasts ---
    last_key = months_data[-1]["key"]
    last_year, last_month = int(last_key.split("-")[0]), int(last_key.split("-")[1])

    forecasts = []
    for step in range(1, forecast_horizon + 1):
        future_month = last_month + step
        future_year = last_year + (future_month - 1) // 12
        future_month = ((future_month - 1) % 12) + 1

        future_x = n - 1 + step
        trend_prediction = slope * future_x + intercept
        seasonal_adj = seasonal_factor.get(future_month, 0.0)
        predicted = max(0, trend_prediction + seasonal_adj)  # no negative bills

        key = f"{future_year}-{future_month:02d}"
        label = f"{month_name[future_month]} {future_year}"
        forecasts.append({
            "key": key,
            "label": label,
            "billed": round(predicted, 2),
            "is_forecast": True,
        })

    # Compute confidence bounds (based on historical residual std dev)
    residuals = []
    for i, val in enumerate(values):
        trend_val = slope * i + intercept + seasonal_factor.get(month_of_year[i], 0.0)
        residuals.append(val - trend_val)

    if len(residuals) > 1:
        mean_r = sum(residuals) / len(residuals)
        std_r = math.sqrt(sum((r - mean_r) ** 2 for r in residuals) / (len(residuals) - 1))
    else:
        std_r = 0.0

    for f in forecasts:
        f["lower"] = round(max(0, f["billed"] - 1.5 * std_r), 2)
        f["upper"] = round(f["billed"] + 1.5 * std_r, 2)

    return forecasts


def analyse_workbooks(uploaded_files):
    monthly = {}
    room_costs = defaultdict(float)
    room_costs_by_month = defaultdict(lambda: defaultdict(float))
    overlap_by_month = defaultdict(float)
    requestors = defaultdict(lambda: {"count": 0, "amount": 0.0})
    totals = {
        "billed": 0.0,
        "rows": 0,
        "matched": 0,
        "zero_charge": 0,
        "unclear": 0,
        "mismatch": 0,
        "missing": 0,
        "unbilled": 0,
        "overlap_savings": 0.0,
    }
    file_summaries = []

    for uploaded in uploaded_files:
        filename = Path(uploaded.filename or "uploaded.xlsx").name
        if filename.startswith("~$") or not filename.lower().endswith(".xlsx"):
            continue

        wb = load_workbook(uploaded, data_only=True)
        if "Reconciliation" not in wb.sheetnames or "All PDF Bookings" not in wb.sheetnames:
            wb.close()
            raise ValueError(f"{filename} is not a reconciled output workbook.")

        recon_rows = get_row_dicts(
            wb["Reconciliation"],
            required_headers=("Date", "Amount", "Status"),
        )
        pdf_rows = get_row_dicts(
            wb["All PDF Bookings"],
            required_headers=("Date", "Location (PDF)", "Expected $"),
        )
        unbilled_rows = []
        if "PDFs Not Billed" in wb.sheetnames:
            unbilled_rows = [
                row for row in get_row_dicts(
                    wb["PDFs Not Billed"],
                    required_headers=("Date", "Location (PDF)"),
                )
                if row.get("Date") and not str(row.get("Date")).startswith("(")
            ]

        month_key, month_label = infer_month_key(filename, recon_rows)
        bucket = monthly.setdefault(month_key, {
            "key": month_key,
            "label": month_label,
            "billed": 0.0,
            "rows": 0,
            "matched": 0,
            "zero_charge": 0,
            "unclear": 0,
            "mismatch": 0,
            "missing": 0,
            "unbilled": 0,
            "overlap_savings": 0.0,
        })

        for row in recon_rows:
            status = normalise_status(row.get("Status"))
            amount = to_float(row.get("Amount"))
            requestor = str(row.get("Requested By") or "Unknown").strip() or "Unknown"

            bucket["rows"] += 1
            bucket["billed"] += amount
            totals["rows"] += 1
            totals["billed"] += amount

            if status == "MATCH":
                bucket["matched"] += 1
                totals["matched"] += 1
            elif status == "ZERO-CHARGE":
                bucket["zero_charge"] += 1
                totals["zero_charge"] += 1
            elif status == "UNCLEAR":
                bucket["unclear"] += 1
                totals["unclear"] += 1
            elif status == "MISMATCH":
                bucket["mismatch"] += 1
                totals["mismatch"] += 1
            elif status == "MISSING":
                bucket["missing"] += 1
                totals["missing"] += 1

            requestors[requestor]["count"] += 1
            requestors[requestor]["amount"] += amount

        for row in pdf_rows:
            rooms = clean_room_labels(row.get("Location (PDF)"))
            expected = to_float(row.get("Expected $"))
            overlap_hours = to_float(row.get("Overlap Hrs"))
            expected_per_room = expected / len(rooms) if rooms else expected
            for room in rooms:
                room_costs[room] += expected_per_room
                room_costs_by_month[month_key][room] += expected_per_room
            savings = overlap_hours * 120.0
            bucket["overlap_savings"] += savings
            totals["overlap_savings"] += savings

        bucket["unbilled"] += len(unbilled_rows)
        totals["unbilled"] += len(unbilled_rows)
        overlap_by_month[month_key] += bucket["overlap_savings"]
        file_summaries.append({"filename": filename, "month": month_label, "rows": len(recon_rows), "pdf_rows": len(pdf_rows)})
        wb.close()

    months = [monthly[key] for key in sorted(monthly)]
    month_count = len(months) or 1
    clean_total = totals["matched"] + totals["zero_charge"]
    totals["match_rate"] = round((clean_total / totals["rows"]) * 100, 1) if totals["rows"] else 0
    totals["avg_monthly_billed"] = round(totals["billed"] / month_count, 2)
    forecast = forecast_billing(months)

    return {
        "files": file_summaries,
        "totals": {key: round(value, 2) if isinstance(value, float) else value for key, value in totals.items()},
        "months": [{key: round(value, 2) if isinstance(value, float) else value for key, value in item.items()} for item in months],
        "forecast": forecast,
        "room_costs": [
            {"room": room, "amount": round(amount, 2)}
            for room, amount in sorted(room_costs.items(), key=lambda item: item[1], reverse=True)
        ],
        "room_costs_by_month": {
            key: [
                {"room": room, "amount": round(amount, 2)}
                for room, amount in sorted(rooms.items(), key=lambda item: item[1], reverse=True)
            ]
            for key, rooms in sorted(room_costs_by_month.items())
        },
        "requestors": [
            {"name": name, "count": data["count"], "amount": round(data["amount"], 2)}
            for name, data in sorted(requestors.items(), key=lambda item: item[1]["count"], reverse=True)[:10]
        ],
    }


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    """Debug endpoint to check if all imports work."""
    status = {"status": "ok", "imports": {}}
    for mod_name in ["flask", "pdfplumber", "openpyxl", "aircon_check_final"]:
        try:
            __import__(mod_name)
            status["imports"][mod_name] = "ok"
        except Exception as e:
            status["imports"][mod_name] = str(e)
            status["status"] = "error"
    return jsonify(status)


@app.route("/upload", methods=["POST"])
def upload():
    """Accept uploaded files, run reconciliation, return summary JSON."""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded."}), 400

    # Create a unique session folder
    session_id = uuid.uuid4().hex[:12]
    session_folder = TEMP_BASE / session_id
    session_folder.mkdir(parents=True, exist_ok=True)

    try:
        # Save uploaded files
        saved_count = {"pdf": 0, "xlsx": 0}
        log_lines = []

        for f in files:
            if not f.filename:
                continue
            # Get just the filename (strip any directory path from folder uploads)
            filename = Path(f.filename).name
            # Skip temp Excel files and already-reconciled outputs
            if filename.startswith("~$") or filename.startswith("Reconciled_"):
                continue
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            if ext == "pdf":
                saved_count["pdf"] += 1
            elif ext == "xlsx":
                saved_count["xlsx"] += 1
            else:
                continue  # Skip non-relevant files

            save_path = session_folder / filename
            f.save(str(save_path))
            log_lines.append(f"  Uploaded: {filename}")

        if saved_count["pdf"] == 0:
            shutil.rmtree(session_folder, ignore_errors=True)
            return jsonify({"error": "No PDF files found in your upload. Please include the Form F PDFs."}), 400
        if saved_count["xlsx"] == 0:
            shutil.rmtree(session_folder, ignore_errors=True)
            return jsonify({"error": "No Excel (.xlsx) file found. Please include the landlord's monthly summary."}), 400

        # Run the original reconcile() function, capturing its stdout
        captured = io.StringIO()
        old_stdout = sys.stdout
        sys.stdout = captured
        try:
            summary_text, output_path = get_reconcile()(str(session_folder))
        finally:
            sys.stdout = old_stdout
        captured_output = captured.getvalue()

        # Parse the text summary into structured data
        result = parse_summary_text(summary_text)

        # Extract output filename
        output_filename = os.path.basename(str(output_path))
        result["output_filename"] = output_filename

        # Build log from captured stdout + summary
        log_lines.append("")
        log_lines.append(captured_output)
        log_lines.append(summary_text)
        result["log_lines"] = log_lines

        # Read the generated output Excel's Summary sheet for accurate values
        try:
            wb_out = load_workbook(str(output_path), data_only=True)
            ws_summary = wb_out["Summary"]
            for row in ws_summary.iter_rows(min_row=2, max_col=2, values_only=True):
                label, value = row
                if label and "Total PDF bookings parsed" in str(label):
                    result["total_pdf_bookings"] = int(value) if value else 0
                elif label and "Total billed amount" in str(label):
                    result["total_billed"] = round(float(value), 2) if value else 0.0
            wb_out.close()
        except Exception:
            # Fallback if we can't read the output Excel
            result.setdefault("total_pdf_bookings", 0)
            result.setdefault("total_billed", 0.0)

        # Store session info for download
        SESSIONS[session_id] = {
            "folder_path": str(session_folder),
            "output_path": str(output_path),
            "output_filename": output_filename,
        }

        result["session_id"] = session_id
        result["download_url"] = f"/download/{session_id}"

        return jsonify(result)

    except Exception as e:
        # Clean up on error
        shutil.rmtree(session_folder, ignore_errors=True)
        return jsonify({"error": str(e)}), 500


@app.route("/analytics", methods=["POST"])
def analytics():
    """Accept reconciled Excel outputs and return analytics JSON."""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No reconciled Excel files uploaded."}), 400

    try:
        result = analyse_workbooks(files)
        if not result["files"]:
            return jsonify({"error": "Please upload one or more Reconciled_*.xlsx files."}), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/download/<session_id>")
def download(session_id):
    """Serve the reconciled Excel file for download."""
    session = SESSIONS.get(session_id)
    if not session:
        return jsonify({"error": "Session expired or not found. Please re-upload your files."}), 404

    output_path = session["output_path"]
    if not os.path.exists(output_path):
        return jsonify({"error": "Output file no longer exists. Please re-upload your files."}), 404

    return send_file(
        output_path,
        as_attachment=True,
        download_name=session["output_filename"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/cleanup/<session_id>", methods=["POST"])
def cleanup(session_id):
    """Clean up temp files for a session."""
    session = SESSIONS.pop(session_id, None)
    if session:
        shutil.rmtree(session["folder_path"], ignore_errors=True)
    return jsonify({"ok": True})


if __name__ == "__main__":
    # Clean up any leftover temp files from previous runs
    if TEMP_BASE.exists():
        for child in TEMP_BASE.iterdir():
            if child.is_dir():
                shutil.rmtree(child, ignore_errors=True)
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("RAILWAY_ENVIRONMENT") is None  # debug only locally
    print(f"\n  Aircon Reconciliation Web App")
    print(f"  Open http://localhost:{port} in your browser\n")
    app.run(debug=debug, host="0.0.0.0", port=port)
