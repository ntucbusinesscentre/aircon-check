"""
Aircon Rental Cross-Check
=========================
Reconciles NTUC Business Centre's aircon-extension Form F PDFs against
the landlord's (CBRE / Mercatus Delta) monthly Excel summary.

Usage (terminal):
    python aircon_check.py                     # prompts for a folder
    python aircon_check.py <folder_path>       # uses that folder

When packaged as aircon_check.exe (PyInstaller --windowed),
double-clicking opens a folder picker dialog.

Output: Reconciled_<month>.xlsx written into the same folder.

Author: built for Alfiq Martindra, NTUC Business Centre.
"""

import os
import re
import sys
import glob
import datetime as dt
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------- Data classes ----------

@dataclass
class PdfBooking:
    """A single booking row extracted from a Form F PDF."""
    location_raw: str          # e.g. "Room 700 (level 7 & 8)"
    location_units: set        # normalised units e.g. {"700", "L7", "L8"}
    date_iso: str              # "2026-04-27"
    date_label: str            # "27 April 2026"
    time_start: str            # "1900"
    time_end: str              # "2200"
    source_pdf: str            # filename
    page_num: int
    # Computed later by analyse_cancellations:
    is_admin: bool = False             # True if this is the L7/L8 admin booking
    overlap_hours: float = 0.0         # hours that overlap with an admin booking
    chargeable_hours: float = 0.0      # hours that should actually be charged
    expected_charge: float = 0.0       # chargeable_hours * rate (× floors for admin)
    overlap_with: list = field(default_factory=list)  # debug: which PDFs overlap with this
    billed_on_bill: bool = False       # True if a landlord row covers this, or it's overlap-cancelled


@dataclass
class ExcelBooking:
    """A single row from the landlord's monthly Excel sheet."""
    row_num: int               # Excel row number (1-indexed)
    date_iso: str
    date_label: str
    time_from: str             # "1900"
    time_to: str               # "2200"
    hours: Optional[float]
    rate: Optional[float]
    additional: Optional[float]
    amount: Optional[float]
    unit_raw: str
    unit_norm_set: set
    requested_by: str
    program_by: str
    remarks: str
    match_status: str = ""     # filled later
    match_pdf: str = ""        # filled later (filename only, for display)
    match_pdf_obj: Optional["PdfBooking"] = None   # filled later (the actual booking)
    match_notes: str = ""      # filled later


# ---------- Location normalisation ----------

LOC_PATTERNS = [
    # Handle "(level 7 & 8)" — allow whitespace anywhere inside the parens
    (re.compile(r"\(\s*level\s*7\s*&\s*8\s*\)", re.I), " L7 L8 "),
    (re.compile(r"\(\s*level\s*7\s*and\s*8\s*\)", re.I), " L7 L8 "),
    # Sometimes the parens have stray spaces and the closing paren is on its own
    (re.compile(r"level\s*7\s*&\s*8", re.I), " L7 L8 "),
    (re.compile(r"\bL\s*7\s*&\s*L\s*8\b", re.I), " L7 L8 "),
    (re.compile(r"\bauditorium\b", re.I), " "),
    (re.compile(r"\b(room|rm|RM)\b\.?", re.I), " "),
    # Mezzanine (= 7M) on level 7 maps to a 7M token so it overlaps with admin units
    (re.compile(r"\bmezzanine\s*,?\s*level\s*7M?\b", re.I), " 7M "),
    (re.compile(r"\bmezzanine\b", re.I), " 7M "),
]

def normalise_units(text):
    """Return a set of normalised unit tokens from a location string.

    Examples:
      "Room 701"                  -> {"701"}
      "Room 700 (level 7 & 8)"    -> {"700", "L7", "L8"}
      "Rm 700/7M/L7/L8"           -> {"700", "7M", "L7", "L8"}
      "Rm 700/702/7M/801/L7/L8"   -> {"700", "702", "7M", "801", "L7", "L8"}
      "Room 7 02" (fragmented)    -> {"702"}
      "Room 700 (le vel 7 & 8)"   -> {"700", "L7", "L8"}  (PDF letter-fragmentation)
      "Mezzanine, Level 7M"       -> {"7M"}
    """
    if not text:
        return set()
    s = str(text)

    # Step 1: de-fragment letter runs inside parentheses.
    # PDFs sometimes split words mid-letter inside (...) e.g. "(le vel 7 & 8)".
    # Collapse all internal whitespace between consecutive letters within parens.
    def _collapse_parens(match):
        inner = match.group(1)
        # Remove whitespace between two letters (handles "le vel" -> "level")
        # but leave whitespace around digits/symbols intact.
        inner = re.sub(r"(?<=[A-Za-z])\s+(?=[A-Za-z])", "", inner)
        return f"({inner})"
    s = re.sub(r"\(([^)]*)\)", _collapse_parens, s)

    # Step 2: apply the high-level patterns
    for pat, rep in LOC_PATTERNS:
        s = pat.sub(rep, s)

    # Step 3: rejoin numeric fragments split by stray whitespace.
    # "7 02" -> "702", "11 04" -> "1104", but DON'T merge "L7 L8" (alpha-prefixed).
    s = re.sub(r"(?<=\d)\s+(?=\d)", "", s)

    # split on slashes, commas, whitespace
    tokens = re.split(r"[\s/,]+", s)
    out = set()
    for tok in tokens:
        tok = tok.strip().upper()
        if not tok:
            continue
        # keep alphanumeric tokens only; ignore stray punctuation
        tok = re.sub(r"[^A-Z0-9]", "", tok)
        if not tok:
            continue
        # Handle compact room labels like "Rm702" / "Room702" the same as
        # "Rm 702" / "Room 702".
        tok = re.sub(r"^(?:ROOM|RM)(?=\d)", "", tok)
        # tokens with a digit are real unit IDs (701, 7M, L7…)
        if any(c.isdigit() for c in tok):
            out.add(tok)
    return out


# Tokens that mean "L7/L8 admin auditorium" — the big shared space spanning two floors.
# Per Alfiq (May 2026): only the auditorium itself counts as admin. 7M (Mezzanine)
# is its own room that sits on L7, and it participates in the overlap-discount logic
# just like 701/801 do — it's NOT part of the admin.
ADMIN_TOKENS = {"L7", "L8"}

def is_admin_booking(unit_set):
    """True if this booking is for the L7/L8 admin auditorium itself.

    A booking is admin only if BOTH L7 and L8 are in its unit set, OR if it
    includes a Rm 700 (the auditorium room number) alongside L7/L8.
    Simpler rule that works: any unit set containing L7 or L8 is admin —
    because the auditorium spans both floors and is always written with at
    least one of those tokens.
    """
    return bool(unit_set & ADMIN_TOKENS)

def is_l7l8_room(unit_set):
    """True if the booking is a non-admin room physically on L7 or L8.

    Includes: 701, 702, 7M (mezzanine on L7), 801, 802, etc.
    Excludes: the admin auditorium itself (handled by is_admin_booking).
    """
    for tok in unit_set:
        # Pure digit room with 7 or 8 as the hundreds digit (701, 801, etc.)
        if tok.isdigit() and len(tok) == 3 and tok[0] in ("7", "8"):
            return True
        # Mezzanine 7M sits on L7 — also eligible for overlap discount
        if tok == "7M":
            return True
    return False


# ---------- Date / time normalisation ----------

MONTHS = {
    "january": "01", "february": "02", "march": "03", "april": "04",
    "may": "05", "june": "06", "july": "07", "august": "08",
    "september": "09", "october": "10", "november": "11", "december": "12",
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "jun": "06", "jul": "07", "aug": "08", "sep": "09",
    "sept": "09", "oct": "10", "nov": "11", "dec": "12",
    # Common PDF text-extraction typo when "February" is split as "Febrau ry".
    "febraury": "02",
    # Common PDF text-extraction typo when "January" is split as "Janau ry".
    "janaury": "01",
}

MONTH_LABELS = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December",
}

def parse_date_text(s):
    """Parse '24 April 2026', '24 April 26', or 'Wednesday, 1 April 2026' -> ('2026-04-24', '24 April 2026').

    Also handles PDF fragmentation artifacts:
      - Year split across text fragments: '01 April 2 026' -> '01 April 2026'
      - Day split: '0 1 April 2026' -> '01 April 2026'
      - Month name split: '17 Apr il 2026' -> '17 April 2026'
    """
    if s is None:
        return None, None
    if isinstance(s, dt.datetime):
        return s.strftime("%Y-%m-%d"), s.strftime("%-d %B %Y") if os.name != "nt" else s.strftime("%#d %B %Y")
    if isinstance(s, dt.date):
        return s.strftime("%Y-%m-%d"), s.strftime("%-d %B %Y") if os.name != "nt" else s.strftime("%#d %B %Y")
    text = str(s).strip()
    # Collapse stray whitespace inside numeric runs: "2 026" -> "2026", "0 1 April" -> "01 April"
    text = re.sub(r"(?<=\d)\s+(?=\d)", "", text)

    # Try the standard pattern first
    m = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{2,4})", text)

    if not m:
        # Try DD/MM/YYYY
        m2 = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text)
        if m2:
            d, mth, y = m2.groups()
            return f"{y}-{int(mth):02d}-{int(d):02d}", f"{int(d)} {list(MONTHS)[int(mth)-1].title()} {y}"

        # Fallback: handle fragmented month names like "Apr il 2026" by trying to
        # find a day-then-letters-then-year pattern, allowing internal whitespace in
        # the letters part, then validating the joined letters spell a real month.
        m3 = re.search(r"(\d{1,2})\s+([A-Za-z][A-Za-z\s]*?[A-Za-z])\s+(\d{2,4})", text)
        if m3:
            day = m3.group(1)
            month_candidate = re.sub(r"\s+", "", m3.group(2)).lower()
            year = m3.group(3)
            if month_candidate in MONTHS:
                if len(year) == 2:
                    year = "20" + year
                mm = MONTHS[month_candidate]
                iso = f"{year}-{mm}-{int(day):02d}"
                label = f"{int(day)} {MONTH_LABELS[mm]} {year}"
                return iso, label
        return None, None

    day, month_word, year = m.groups()
    mm = MONTHS.get(month_word.lower())
    if not mm:
        return None, None
    if len(year) == 2:
        year = "20" + year
    iso = f"{year}-{mm}-{int(day):02d}"
    label = f"{int(day)} {MONTH_LABELS[mm]} {year}"
    return iso, label


def normalise_time(s):
    """'1900hrs' / '19 00 hrs' / '1900' / 1900 / '190 0hrs' -> '1900'.
    Returns '' if can't parse."""
    if s is None:
        return ""
    text = str(s).strip()
    # strip everything except digits
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    # if it looks like 3 digits, pad to 4
    if len(digits) == 3:
        digits = "0" + digits
    if len(digits) == 4:
        return digits
    if len(digits) > 4:
        # take first 4
        return digits[:4]
    return digits.zfill(4)


def parse_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    # Strip $, commas, ALL whitespace including non-breaking space (\xa0) and other unicode
    s = str(v)
    # Remove currency symbols, commas, and all whitespace variants
    for ch in ("$", ",", "\xa0", "\u202f", "\u2009", "\u00a0"):
        s = s.replace(ch, "")
    s = re.sub(r"\s", "", s)  # remove any remaining whitespace
    if s in ("", "-", "-", "—", "–"):  # blank, dash, em-dash, en-dash
        return 0.0
    try:
        return float(s)
    except ValueError:
        return None


def compute_hours(time_start, time_end):
    """Compute hour delta between two HHMM strings. Returns float or None."""
    if not time_start or not time_end:
        return None
    try:
        ts = int(time_start[:2]) + int(time_start[2:]) / 60.0
        te = int(time_end[:2]) + int(time_end[2:]) / 60.0
        if te < ts:
            te += 24  # overnight
        return round(te - ts, 2)
    except (ValueError, IndexError):
        return None


# ---------- PDF parsing ----------

def extract_pdf_bookings(pdf_path):
    """Extract all booking rows from a Form F PDF.

    Returns a list of PdfBooking.
    """
    bookings = []
    fname = os.path.basename(pdf_path)
    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            page_num = page_idx + 1
            tables = page.extract_tables()
            for table in tables:
                # the booking rows live below the row that has "Location | Date | Time (Start) | ... | Time (End)"
                in_section = False
                for row in table:
                    if not row:
                        continue
                    # clean cells: convert None to "", strip
                    cells = [("" if c is None else str(c)).strip() for c in row]
                    joined = " | ".join(cells).lower()

                    if "location" in joined and "date" in joined and "time" in joined:
                        in_section = True
                        continue
                    if "payment mode" in joined or "billing details" in joined:
                        in_section = False
                        continue

                    if not in_section:
                        continue

                    booking = parse_booking_row(cells, fname, page_num)
                    if booking:
                        bookings.append(booking)
    return bookings


def parse_booking_row(cells, fname, page_num):
    """Try to extract location/date/time from a list of cell strings."""
    # filter out empty cells but keep order
    nonempty = [c for c in cells if c]
    if len(nonempty) < 3:
        return None

    # find the cell containing a date like "24 April 2026"
    date_iso, date_label = None, None
    date_idx = -1
    for i, c in enumerate(nonempty):
        iso, lbl = parse_date_text(c)
        if iso:
            date_iso, date_label = iso, lbl
            date_idx = i
            break
    if date_iso is None:
        return None

    # location is the cell before the date cell
    if date_idx == 0:
        return None
    location_raw = nonempty[date_idx - 1]

    # the time cells are after the date cell
    time_cells = nonempty[date_idx + 1:]
    times = []
    for c in time_cells:
        t = normalise_time(c)
        if t:
            times.append(t)
    if len(times) < 2:
        return None

    return PdfBooking(
        location_raw=location_raw,
        location_units=normalise_units(location_raw),
        date_iso=date_iso,
        date_label=date_label,
        time_start=times[0],
        time_end=times[1],
        source_pdf=fname,
        page_num=page_num,
    )


# ---------- Excel parsing ----------

EXCEL_HEADER_KEYS = {
    "date": re.compile(r"^date", re.I),
    "from": re.compile(r"^from", re.I),
    "to":   re.compile(r"^to\b", re.I),
    "hours": re.compile(r"hours", re.I),
    "rate": re.compile(r"rate", re.I),
    "additional": re.compile(r"additional", re.I),
    "amount": re.compile(r"amount", re.I),
    "unit": re.compile(r"^unit", re.I),
    "requested_by": re.compile(r"requested", re.I),
    "program_by": re.compile(r"program", re.I),
    "remarks": re.compile(r"remark", re.I),
}


def find_header_row(ws):
    """Find the header row index (1-indexed) and a dict of {field: col_idx}.

    Also handles two-row headers (where row 1 has 'No. of' and row 2 has 'Hours',
    etc.) by combining the two rows when row 1 alone is incomplete.
    """
    for row_idx in range(1, min(20, ws.max_row + 1)):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        cell_strs = [("" if v is None else str(v)).strip() for v in cells]
        lower = [s.lower() for s in cell_strs]
        # must look like a header: has "date", "from", "amount"
        has_date = any(s.startswith("date") for s in lower)
        has_from = any(s.startswith("from") for s in lower)
        has_amount = any("amount" in s for s in lower)
        if has_date and has_from and has_amount:
            # build col map from this row
            col_map = {}
            for ci, s in enumerate(cell_strs):
                for key, pat in EXCEL_HEADER_KEYS.items():
                    if pat.search(s):
                        if key not in col_map:
                            col_map[key] = ci + 1  # 1-indexed for openpyxl

            # If the next row also looks like a header continuation (no real data,
            # just sub-headers like "(hrs)", "Hours", "Hour ($)"), merge its info.
            # This catches landlord sheets where the column headers wrap onto 2 rows.
            next_row_idx = row_idx + 1
            if next_row_idx <= ws.max_row:
                next_cells = [ws.cell(row=next_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
                next_strs = [("" if v is None else str(v)).strip() for v in next_cells]
                next_lower = [s.lower() for s in next_strs]
                # heuristic: this is a header-continuation row if every non-empty cell
                # is short text (no real date/number that would indicate it's a data row)
                non_empty = [s for s in next_strs if s]
                looks_like_subheader = (len(non_empty) > 0
                    and not any(re.search(r"\d{4}", s) for s in next_strs)   # no years
                    and not any(re.search(r"^\s*\d{3,4}\s*$", s) for s in next_strs)  # no times
                    and all(len(s) < 25 for s in non_empty))
                if looks_like_subheader:
                    # Fill in any missing fields by checking row 2
                    for ci, s in enumerate(next_strs):
                        for key, pat in EXCEL_HEADER_KEYS.items():
                            if pat.search(s) and key not in col_map:
                                col_map[key] = ci + 1
                    # data rows actually start AFTER the sub-header
                    return next_row_idx, col_map

            return row_idx, col_map
    return None, None


def extract_excel_bookings(xlsx_path):
    """Returns (workbook, sheet, header_row, col_map, list_of_ExcelBooking)."""
    wb = load_workbook(xlsx_path, data_only=True)
    # use the first sheet with a recognisable header
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, col_map = find_header_row(ws)
        if header_row is not None:
            bookings = []
            for r in range(header_row + 1, ws.max_row + 1):
                date_val = ws.cell(row=r, column=col_map.get("date", 1)).value
                if date_val is None or str(date_val).strip() == "":
                    continue
                date_str = str(date_val).strip()
                # skip subtotal / grand-total rows
                if re.search(r"sub.?total|grand.?total", date_str, re.I):
                    continue
                date_iso, date_label = parse_date_text(date_val)
                if date_iso is None:
                    continue

                def gv(key, default=""):
                    if key in col_map:
                        v = ws.cell(row=r, column=col_map[key]).value
                        return v if v is not None else default
                    return default

                unit_raw = str(gv("unit", "")).strip()
                time_from = normalise_time(gv("from"))
                time_to = normalise_time(gv("to"))
                hours_val = parse_number(gv("hours"))
                # If the landlord's sheet stores Hours as a formula that hasn't been
                # cached (or the cell is blank), fall back to computing it from times.
                if hours_val is None or hours_val == 0:
                    computed = compute_hours(time_from, time_to)
                    if computed is not None:
                        hours_val = computed

                booking = ExcelBooking(
                    row_num=r,
                    date_iso=date_iso,
                    date_label=date_label,
                    time_from=time_from,
                    time_to=time_to,
                    hours=hours_val,
                    rate=parse_number(gv("rate")),
                    additional=parse_number(gv("additional")),
                    amount=parse_number(gv("amount")),
                    unit_raw=unit_raw,
                    unit_norm_set=normalise_units(unit_raw),
                    requested_by=str(gv("requested_by", "")).strip(),
                    program_by=str(gv("program_by", "")).strip(),
                    remarks=str(gv("remarks", "")).strip(),
                )
                bookings.append(booking)
            return wb, ws, header_row, col_map, bookings
    raise RuntimeError(f"Could not find a header row in {xlsx_path}. "
                       f"Expected columns: Date, From, To, Hours, Rate, Amount, Unit, ...")


# ---------- Overlap & cancellation analysis ----------

RATE_PER_HOUR_FLOOR = 120.0
ADMIN_FLOOR_COUNT = 2  # admin auditorium spans L7 + L8

def time_to_minutes(hhmm):
    """'1900' -> 1140 (minutes since midnight). Returns None if invalid."""
    if not hhmm or len(hhmm) < 3:
        return None
    try:
        h = int(hhmm[:2])
        m = int(hhmm[2:])
        return h * 60 + m
    except (ValueError, IndexError):
        return None

def overlap_minutes(start_a, end_a, start_b, end_b):
    """Return the overlap in minutes between two HHMM time windows on the same day."""
    a1, a2 = time_to_minutes(start_a), time_to_minutes(end_a)
    b1, b2 = time_to_minutes(start_b), time_to_minutes(end_b)
    if None in (a1, a2, b1, b2):
        return 0
    # handle overnight (e.g. 2300 -> 0100): if end < start, add 24h
    if a2 < a1: a2 += 24 * 60
    if b2 < b1: b2 += 24 * 60
    lo, hi = max(a1, b1), min(a2, b2)
    return max(0, hi - lo)


def _minutes_to_hhmm(mins):
    """Convert minutes since midnight back to 'HHMM' string. Wraps over 24h."""
    mins = mins % (24 * 60)
    return f"{mins // 60:02d}{mins % 60:02d}"


def _subtract_admin_windows(room_start, room_end, admin_windows):
    """Subtract a list of (start, end) admin time windows from the room's window.

    Returns a list of (start, end) tuples in HHMM format representing the
    *remaining* (chargeable) sub-windows. Used to verify whether a shortened
    bill time-window represents a correct cancellation.

    Example: room 1900-2100, admin 1900-2000 → returns [('2000', '2100')]
    """
    rs = time_to_minutes(room_start)
    re_ = time_to_minutes(room_end)
    if rs is None or re_ is None:
        return []
    if re_ < rs:
        re_ += 24 * 60

    # Start with the full room window as the only "kept" segment
    kept = [(rs, re_)]
    for a_start, a_end in admin_windows:
        as_ = time_to_minutes(a_start)
        ae_ = time_to_minutes(a_end)
        if as_ is None or ae_ is None:
            continue
        if ae_ < as_:
            ae_ += 24 * 60
        # Subtract (as_, ae_) from every kept segment
        new_kept = []
        for ks, ke in kept:
            # No overlap
            if ae_ <= ks or as_ >= ke:
                new_kept.append((ks, ke))
                continue
            # Left piece survives
            if as_ > ks:
                new_kept.append((ks, as_))
            # Right piece survives
            if ae_ < ke:
                new_kept.append((ae_, ke))
        kept = new_kept

    return [(_minutes_to_hhmm(s), _minutes_to_hhmm(e)) for s, e in kept]


def analyse_cancellations(pdf_bookings):
    """Mark each PDF booking with its admin/overlap status and expected charge.

    Business rule (per Alfiq, May 2026):
      - L7/L8 admin auditorium = always charged in full at $120/hr × 2 floors.
        Appears on the bill as TWO duplicate rows (one per floor).
      - Rooms physically on L7 or L8 (701, 702, 801, 7M etc): if they overlap
        in time with an admin booking on the SAME DATE, the overlap hours are
        free. Only the non-overlap (extra) hours are charged at $120/hr.
      - Rooms NOT on L7/L8 (903, 1104 etc): charged normally, no cancellation.
    """
    # Group bookings by date for efficiency
    by_date = {}
    for p in pdf_bookings:
        by_date.setdefault(p.date_iso, []).append(p)

    for date_iso, day_bookings in by_date.items():
        # find admin bookings on this date
        admin_bookings = [p for p in day_bookings if is_admin_booking(p.location_units)]
        for p in day_bookings:
            total_h = compute_hours(p.time_start, p.time_end) or 0.0
            p.is_admin = is_admin_booking(p.location_units)

            if p.is_admin:
                # Admin always charged in full, on both floors
                p.overlap_hours = 0.0
                p.chargeable_hours = total_h
                p.expected_charge = round(total_h * RATE_PER_HOUR_FLOOR * ADMIN_FLOOR_COUNT, 2)
                continue

            # Non-admin room. Does it overlap with any admin booking today?
            if not admin_bookings or not is_l7l8_room(p.location_units):
                # Either no admin booking today, or this room isn't on L7/L8 — full charge
                p.overlap_hours = 0.0
                p.chargeable_hours = total_h
                p.expected_charge = round(total_h * RATE_PER_HOUR_FLOOR, 2)
                continue

            # Compute total overlap minutes with admin bookings (union of admin windows)
            # Simple approach: take the maximum overlap with any single admin booking,
            # which is correct as long as admin bookings on the same day don't themselves
            # overlap. If they do, we sum overlaps but cap at total room time.
            overlap_min_total = 0
            overlap_partners = []
            for a in admin_bookings:
                om = overlap_minutes(p.time_start, p.time_end, a.time_start, a.time_end)
                if om > 0:
                    overlap_min_total += om
                    overlap_partners.append(f"{a.location_raw} {a.time_start}-{a.time_end}")
            overlap_h = round(overlap_min_total / 60.0, 2)
            # cap at total room hours (in case of multiple overlapping admins)
            overlap_h = min(overlap_h, total_h)
            p.overlap_hours = overlap_h
            p.chargeable_hours = round(total_h - overlap_h, 2)
            p.expected_charge = round(p.chargeable_hours * RATE_PER_HOUR_FLOOR, 2)
            p.overlap_with = overlap_partners


# ---------- Matching ----------

def units_overlap(a, b):
    """Do two unit-sets overlap?"""
    if not a or not b:
        return False
    return bool(a & b)


def match_booking(excel_row, pdf_bookings):
    """Find best PDF match for an excel row.

    Returns (status, pdf_booking_or_None, note).
    Status in: MATCH, ZERO-CHARGE, MISMATCH, MISSING, UNCLEAR.

    Notes on admin (L7/L8) billing:
      Admin bookings appear on the bill as TWO duplicate rows (one per floor).
      Each row's expected amount = pdf_hours × $120 (one floor's worth).
      So when comparing, divide the PDF's total expected_charge by 2 for admins.

    Notes on room overlap cancellation:
      Rooms on L7/L8 that overlap admin bookings are charged only for non-overlap
      hours. The script uses chargeable_hours (not total hours) for expected amount.
    """
    # First: date matches
    same_date = [p for p in pdf_bookings if p.date_iso == excel_row.date_iso]
    if not same_date:
        return "MISSING", None, "No PDF uploaded for this date (upload the form if you have it)"

    # Within same date, try to match by time AND location overlap
    exact = [p for p in same_date
             if p.time_start == excel_row.time_from
             and p.time_end == excel_row.time_to
             and units_overlap(p.location_units, excel_row.unit_norm_set)]
    if exact:
        chosen = exact[0]
        pdf_hours = compute_hours(chosen.time_start, chosen.time_end)
        notes = []

        # Determine the expected billed amount for THIS particular line
        # Admin = bill has 2 rows, each charging hours × $120 (per floor)
        # Room  = bill has 1 row, charging chargeable_hours × $120 (after overlap deduction)
        if chosen.is_admin:
            expected = round((pdf_hours or 0) * RATE_PER_HOUR_FLOOR, 2)
            line_label = "admin (per floor)"
        else:
            expected = round(chosen.chargeable_hours * RATE_PER_HOUR_FLOOR, 2)
            line_label = "room"
            if chosen.overlap_hours > 0:
                line_label = f"room after {chosen.overlap_hours:g}h overlap deduction"

        # Hours discrepancy
        if (excel_row.hours is not None and pdf_hours is not None
                and abs(excel_row.hours - pdf_hours) > 0.01):
            notes.append(f"Hours mismatch: PDF says {pdf_hours:g}h, bill claims {excel_row.hours:g}h")

        # Zero charge — flag if our expectation is non-zero
        if excel_row.amount is not None and excel_row.amount == 0:
            if expected > 0:
                return "ZERO-CHARGE", chosen, (f"Bill shows $0 but expected ${expected:.2f} "
                                               f"({line_label}). Check L7/L8 cancellation rule.")
            return "ZERO-CHARGE", chosen, "Bill shows $0 (consistent with full cancellation)"

        # Amount discrepancy
        if (excel_row.amount is not None and expected > 0
                and abs(expected - excel_row.amount) > 0.5):
            notes.append(f"Amount mismatch: expected ${expected:.2f} ({line_label}), "
                         f"bill shows ${excel_row.amount:.2f}")

        # Add a contextual note for non-trivial cases even when amounts match
        if not notes and chosen.overlap_hours > 0 and not chosen.is_admin:
            partners = ", ".join(chosen.overlap_with) if chosen.overlap_with else "admin booking"
            notes.append(f"{chosen.overlap_hours:g}h overlap with {partners} — "
                         f"only {chosen.chargeable_hours:g}h charged")

        if any("mismatch" in n.lower() for n in notes):
            return "MISMATCH", chosen, "; ".join(notes)
        return "MATCH", chosen, "; ".join(notes)

    # Same date + same time, but no location overlap
    time_only = [p for p in same_date
                 if p.time_start == excel_row.time_from
                 and p.time_end == excel_row.time_to]
    if time_only:
        chosen = time_only[0]
        other_pdfs = ", ".join(set(p.location_raw for p in time_only))
        return "UNCLEAR", chosen, (f"⚠ Possible billing error: bill charges '{excel_row.unit_raw}' "
                                    f"at {excel_row.time_from}-{excel_row.time_to}, "
                                    f"but PDF on this date only shows {other_pdfs} at this time")

    # Same date, overlapping units, but different times
    loc_only = [p for p in same_date if units_overlap(p.location_units, excel_row.unit_norm_set)]
    if loc_only:
        chosen = loc_only[0]
        # Special case: the landlord may charge only the non-overlap portion of a room
        # booking when it overlaps with an admin booking. The bill's time window will
        # then be a sub-window of the PDF's. Verify by computing what the post-cancellation
        # time window would be.
        if (not chosen.is_admin and chosen.overlap_hours > 0
                and chosen.chargeable_hours > 0):
            # The "kept" portion of the room's time = whatever doesn't overlap admin.
            # Compute it by subtracting admin windows from the room's full window.
            kept_windows = _subtract_admin_windows(
                chosen.time_start, chosen.time_end,
                [(p.time_start, p.time_end) for p in same_date if p.is_admin]
            )
            bill_window = (excel_row.time_from, excel_row.time_to)
            if bill_window in kept_windows:
                # The bill is correctly showing the post-cancellation time slice.
                amount_ok = (excel_row.amount is not None
                             and abs(excel_row.amount - chosen.chargeable_hours * RATE_PER_HOUR_FLOOR) < 0.5)
                if amount_ok:
                    partners = ", ".join(chosen.overlap_with) or "admin booking"
                    return "MATCH", chosen, (f"Bill shows post-cancellation slice "
                                              f"{excel_row.time_from}-{excel_row.time_to} only; "
                                              f"{chosen.overlap_hours:g}h cancelled by {partners}")
        return "UNCLEAR", chosen, (f"Date+unit match, but times differ "
                                   f"(PDF: {chosen.time_start}-{chosen.time_end}; "
                                   f"bill: {excel_row.time_from}-{excel_row.time_to})")

    # Same date, but the unit on the bill is nowhere in any PDF for this day
    # → this is more suspicious than "no PDF for this date" because PDFs DO exist
    if same_date:
        pdf_units_on_day = set()
        for p in same_date:
            pdf_units_on_day.update(p.location_units)
        return "MISSING", None, (f"⚠ Possible billing error: PDFs exist for this date but "
                                  f"none mentions '{excel_row.unit_raw}'. "
                                  f"PDFs on {excel_row.date_label} cover: "
                                  f"{', '.join(sorted(pdf_units_on_day))}")

    return "MISSING", None, "No PDF uploaded for this date (upload the form if you have it)"


# ---------- Output workbook ----------

# Style constants — colours match the masthead aesthetic of the original sheet
FILL_HEADER = PatternFill("solid", fgColor="1A1614")
FILL_OK = PatternFill("solid", fgColor="D4E4D8")
FILL_ZERO = PatternFill("solid", fgColor="F0E2B8")
FILL_FLAG = PatternFill("solid", fgColor="F0D0C8")
FILL_UNCLEAR = PatternFill("solid", fgColor="F8E5C8")
FILL_GREY = PatternFill("solid", fgColor="EDEDED")
# Distinct colours for the All PDF Bookings tab:
FILL_ADMIN = PatternFill("solid", fgColor="BFD7ED")      # blue   = Room 700 / auditorium bookings
FILL_CANCELLED = PatternFill("solid", fgColor="E2D4F0")  # purple = rooms cancelled / reduced by overlap
FILL_RED = PatternFill("solid", fgColor="F0C0B8")        # red    = not on bill / mismatch

THIN = Side(border_style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_output(xlsx_path, excel_bookings, pdf_bookings, used_pdf_ids, output_path):
    """Build the reconciled workbook."""
    wb = Workbook()

    # --- Sheet 1: Reconciliation ---
    ws = wb.active
    ws.title = "Reconciliation"

    headers = [
        "Date", "From", "To", "Hours", "Rate", "Additional", "Amount",
        "Unit (Landlord)", "Requested By", "Program By", "Remarks",
        "Status", "Expected $", "PDF Source", "PDF Location", "PDF Time", "Notes",
    ]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = FILL_HEADER
        c.font = Font(bold=True, color="F4EDE0", name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX

    for ri, eb in enumerate(excel_bookings, start=2):
        match_pdf = eb.match_pdf_obj

        # Compute expected $ for this landlord line
        if match_pdf is not None:
            if match_pdf.is_admin:
                # admin: each row charges 1 floor's worth
                expected = round((compute_hours(match_pdf.time_start, match_pdf.time_end) or 0)
                                  * RATE_PER_HOUR_FLOOR, 2)
            else:
                expected = match_pdf.expected_charge
        else:
            expected = None

        row_vals = [
            eb.date_label,
            eb.time_from,
            eb.time_to,
            eb.hours,
            eb.rate,
            eb.additional,
            eb.amount,
            eb.unit_raw,
            eb.requested_by,
            eb.program_by,
            eb.remarks,
            eb.match_status,
            expected,
            eb.match_pdf,
            match_pdf.location_raw if match_pdf else "",
            f"{match_pdf.time_start}-{match_pdf.time_end}" if match_pdf else "",
            eb.match_notes,
        ]
        for ci, v in enumerate(row_vals, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(ci >= 11))
            c.border = BOX
            if ci == 4:  # hours
                c.number_format = "0.##"
            if ci in (5, 6, 7, 13):  # rate, additional, amount, expected
                c.number_format = '"$"#,##0.00'

        # row-level colour by status (unified scheme across all tabs):
        #   white  = clean match (no fill)
        #   purple = cancellation-related (zero-charge from overlap)
        #   red    = problem needing review (missing / mismatch / unclear)
        fill = None
        if eb.match_status == "MATCH":
            fill = None
        elif eb.match_status == "ZERO-CHARGE":
            fill = FILL_CANCELLED
        elif eb.match_status == "MISSING":
            fill = FILL_RED
        elif eb.match_status == "MISMATCH":
            fill = FILL_RED
        elif eb.match_status == "UNCLEAR":
            fill = FILL_RED

        if fill:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    # column widths (17 columns now, including Expected $)
    widths = [22, 8, 8, 7, 10, 12, 12, 28, 14, 14, 28, 16, 12, 38, 24, 14, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # --- Sheet 2: Unused PDF entries (PDF booking with no landlord row) ---
    ws2 = wb.create_sheet("PDFs Not Billed")
    headers2 = ["Date", "Location (PDF)", "Time Start", "Time End", "Source PDF", "Page"]
    for ci, h in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = FILL_HEADER
        c.font = Font(bold=True, color="F4EDE0", name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX

    unused = [p for p in pdf_bookings if not p.billed_on_bill]
    for ri, p in enumerate(unused, start=2):
        vals = [p.date_label, p.location_raw, p.time_start, p.time_end, p.source_pdf, p.page_num]
        for ci, v in enumerate(vals, start=1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10)
            c.border = BOX
            c.fill = FILL_RED  # red = problem (not on bill)
    if not unused:
        ws2.cell(row=2, column=1, value="(none — every PDF booking appears on the bill)").font = Font(name="Arial", italic=True)
    for i, w in enumerate([22, 32, 12, 12, 48, 8], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # --- Sheet 3: All PDF Bookings (raw extraction for inspection) ---
    ws_pdfs = wb.create_sheet("All PDF Bookings")

    # Colour legend across the top (row 1)
    legend = [
        ("Colour key:", None),
        ("  Room 700 / auditorium", FILL_ADMIN),
        ("  Cancelled / reduced by overlap", FILL_CANCELLED),
        ("  Not on bill / mismatch", FILL_RED),
        ("  Normal room booking", None),
    ]
    lc = 1
    for label, fill in legend:
        cell = ws_pdfs.cell(row=1, column=lc, value=label)
        cell.font = Font(name="Arial", size=9, bold=(fill is None and label.startswith("Colour")))
        if fill:
            cell.fill = fill
        lc += 1

    pdf_headers = [
        "Date", "Location (PDF)", "Time Start", "Time End", "Hours",
        "Type", "Overlap Hrs", "Chargeable Hrs", "Expected $",
        "Billed?", "Source PDF", "Page",
    ]
    for ci, h in enumerate(pdf_headers, start=1):
        c = ws_pdfs.cell(row=2, column=ci, value=h)
        c.fill = FILL_HEADER
        c.font = Font(bold=True, color="F4EDE0", name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX

    # Sort by date, then by source PDF, then by time
    sorted_pdfs = sorted(pdf_bookings,
                         key=lambda p: (p.date_iso or "", p.source_pdf, p.time_start))
    for ri, p in enumerate(sorted_pdfs, start=3):
        is_billed = p.billed_on_bill
        billed = "Yes" if is_billed else "No — see 'PDFs Not Billed'"
        booking_type = "Admin (L7/L8)" if p.is_admin else "Room"
        total_h = compute_hours(p.time_start, p.time_end) or 0
        vals = [
            p.date_label,
            p.location_raw,
            p.time_start,
            p.time_end,
            total_h,
            booking_type,
            p.overlap_hours if p.overlap_hours > 0 else "",
            p.chargeable_hours if not p.is_admin else "",
            p.expected_charge,
            billed,
            p.source_pdf,
            p.page_num,
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws_pdfs.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 11))
            c.border = BOX
            if ci in (5, 7, 8):
                c.number_format = "0.##"
            if ci == 9:
                c.number_format = '"$"#,##0.00'
        # Colour code: red = not on bill / mismatch; blue = Room 700 admin;
        # purple = room cancelled or reduced by overlap; white = normal room booking.
        if not is_billed:
            fill = FILL_RED
        elif p.is_admin:
            fill = FILL_ADMIN
        elif p.overlap_hours > 0:
            fill = FILL_CANCELLED
        else:
            fill = None
        if fill:
            for ci in range(1, len(pdf_headers) + 1):
                ws_pdfs.cell(row=ri, column=ci).fill = fill

    if not pdf_bookings:
        ws_pdfs.cell(row=3, column=1, value="(no PDF bookings parsed)").font = Font(name="Arial", italic=True)
    for i, w in enumerate([22, 30, 11, 11, 7, 16, 11, 13, 12, 28, 48, 6], start=1):
        ws_pdfs.column_dimensions[get_column_letter(i)].width = w
    ws_pdfs.freeze_panes = "A3"
    # Auto-filter on the header row (row 2) through the last data row
    ws_pdfs.auto_filter.ref = f"A2:{get_column_letter(len(pdf_headers))}{max(2, ws_pdfs.max_row)}"

    # --- Sheet 5: Summary ---
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Aircon Rental Reconciliation Summary"
    ws3["A1"].font = Font(bold=True, size=14, name="Arial")
    ws3.merge_cells("A1:D1")

    counts = {"MATCH": 0, "ZERO-CHARGE": 0, "UNCLEAR": 0, "MISMATCH": 0, "MISSING": 0}
    for eb in excel_bookings:
        counts[eb.match_status] = counts.get(eb.match_status, 0) + 1

    rows = [
        ["", "", "", ""],
        ["Total landlord rows", len(excel_bookings), "", ""],
        ["", "", "", ""],
        ["Matched cleanly",          counts.get("MATCH", 0),       "OK",                     ""],
        ["Zero-charge (review)",     counts.get("ZERO-CHARGE", 0), "Likely L7/L8 cancellation", ""],
        ["Unclear (review)",         counts.get("UNCLEAR", 0),     "Date matches but time or unit differs", ""],
        ["Amount mismatch",          counts.get("MISMATCH", 0),    "Expected ≠ billed amount", ""],
        ["Missing from PDFs",        counts.get("MISSING", 0),     "No matching PDF booking", ""],
        ["PDF entries not billed",   len(unused),                  "See sheet 'PDFs Not Billed'", ""],
        ["Total PDF bookings parsed", len(pdf_bookings), "See sheet 'All PDF Bookings'", ""],
        ["", "", "", ""],
        ["Total billed amount", sum((eb.amount or 0) for eb in excel_bookings), "", ""],
    ]
    for ri, r in enumerate(rows, start=2):
        for ci, v in enumerate(r, start=1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=10, bold=(ci == 1))
            c.alignment = Alignment(vertical="center")
            if ci == 2 and isinstance(v, (int, float)) and r[0] == "Total billed amount":
                c.number_format = '"$"#,##0.00'

    # Tint the label cell of each category row to match the colour used on the tabs,
    # so the Summary doubles as a colour legend consistent with every other sheet.
    summary_fills = {
        "Zero-charge (review)": FILL_CANCELLED,
        "Unclear (review)": FILL_RED,
        "Amount mismatch": FILL_RED,
        "Missing from PDFs": FILL_RED,
        "PDF entries not billed": FILL_RED,
    }
    for ri, r in enumerate(rows, start=2):
        if r[0] in summary_fills:
            ws3.cell(row=ri, column=1).fill = summary_fills[r[0]]

    for i, w in enumerate([28, 14, 50, 8], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)


# ---------- Folder picker ----------

def pick_folder():
    """Open a folder picker. Falls back to terminal prompt if tkinter is unavailable."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        folder = filedialog.askdirectory(title="Choose the folder with PDFs + landlord's Excel")
        root.destroy()
        return folder
    except Exception:
        return input("Enter the folder path containing PDFs and landlord's Excel: ").strip()


def show_message(title, message, is_error=False):
    """Show a popup if Tk is available, else print to console."""
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        if is_error:
            messagebox.showerror(title, message)
        else:
            messagebox.showinfo(title, message)
        root.destroy()
    except Exception:
        prefix = "ERROR: " if is_error else ""
        print(f"\n{prefix}{title}\n{message}\n")


# ---------- Main ----------

def reconcile(folder):
    folder = Path(folder)
    if not folder.exists() or not folder.is_dir():
        raise RuntimeError(f"Folder not found: {folder}")

    pdf_paths = sorted(folder.glob("*.pdf"))
    xlsx_paths = [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")
                  and not p.name.startswith("Reconciled_")]

    if not pdf_paths:
        raise RuntimeError(f"No PDF files found in {folder}")
    if not xlsx_paths:
        raise RuntimeError(f"No .xlsx file found in {folder} (excluding files starting with 'Reconciled_').")
    if len(xlsx_paths) > 1:
        # use the most recently modified — but warn loudly
        xlsx_paths.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        print(f"  ⚠  Multiple Excel files found. Using the most recent:")
        for p in xlsx_paths:
            marker = "→" if p == xlsx_paths[0] else " "
            print(f"     {marker} {p.name}")

    xlsx_path = xlsx_paths[0]
    print(f"  Excel: {xlsx_path.name}")
    print(f"  PDFs ({len(pdf_paths)}):")
    for p in pdf_paths:
        print(f"    - {p.name}")

    # parse PDFs
    pdf_bookings = []
    for p in pdf_paths:
        bookings = extract_pdf_bookings(p)
        pdf_bookings.extend(bookings)
        print(f"  · parsed {len(bookings):2d} bookings from {p.name}")

    if not pdf_bookings:
        raise RuntimeError("No bookings could be extracted from the PDFs.")

    # Analyse L7/L8 overlap cancellations
    analyse_cancellations(pdf_bookings)
    overlap_count = sum(1 for p in pdf_bookings if p.overlap_hours > 0)
    if overlap_count:
        print(f"  · detected {overlap_count} L7/L8 overlap-cancellation(s)")

    # parse Excel
    wb_in, ws_in, header_row, col_map, excel_bookings = extract_excel_bookings(xlsx_path)
    print(f"  · parsed {len(excel_bookings)} rows from {xlsx_path.name}")

    # match
    used_pdf_ids = set()
    for eb in excel_bookings:
        status, pdf_match, note = match_booking(eb, pdf_bookings)
        eb.match_status = status
        eb.match_notes = note
        if pdf_match:
            eb.match_pdf = pdf_match.source_pdf
            eb.match_pdf_obj = pdf_match
            used_pdf_ids.add(id(pdf_match))  # unique by object id

    # convert object-id set to index set
    used_idx = {i for i, p in enumerate(pdf_bookings) if id(p) in used_pdf_ids}

    # Determine "is this PDF booking accounted for?" independently of the strict
    # 1-to-1 match. A booking counts as billed/accounted-for if EITHER:
    #   (a) any landlord row on the same date overlaps it in time AND shares a unit
    #       token (covers auditorium days where one auditorium PDF is billed as two
    #       rows, and overlapping rooms are absorbed into the auditorium charge), OR
    #   (b) it was fully or partly cancelled by an overlap (appears in the L7-L8
    #       Cancellations tab) — these are intentionally not charged, so not a problem.
    for p in pdf_bookings:
        covered = False
        # (b) overlap-cancelled rooms are accounted for
        if p.overlap_hours > 0:
            covered = True
        else:
            # (a) covered by a landlord row on the same date+time+unit
            for eb in excel_bookings:
                if eb.date_iso != p.date_iso:
                    continue
                if not units_overlap(p.location_units, eb.unit_norm_set):
                    continue
                if overlap_minutes(p.time_start, p.time_end, eb.time_from, eb.time_to) > 0:
                    covered = True
                    break
        p.billed_on_bill = covered

    # decide output filename based on the first booking's month
    if excel_bookings:
        # use the month that appears most often
        from collections import Counter
        month_counts = Counter(b.date_iso[:7] for b in excel_bookings if b.date_iso)
        month_iso = month_counts.most_common(1)[0][0]  # e.g. "2026-04"
        y, m = month_iso.split("-")
        month_name = list(MONTHS)[int(m)-1].title()
        out_name = f"Reconciled_{month_name}_{y}.xlsx"
    else:
        out_name = "Reconciled.xlsx"

    out_path = folder / out_name
    # Try the main path; if locked (file open in Excel), fall back to a timestamped name
    try:
        write_output(xlsx_path, excel_bookings, pdf_bookings, used_idx, out_path)
    except PermissionError:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_name = out_name.replace(".xlsx", f"_{ts}.xlsx")
        fallback_path = folder / fallback_name
        print(f"  ⚠  {out_name} is locked (probably open in Excel).")
        print(f"     Saving as {fallback_name} instead.")
        write_output(xlsx_path, excel_bookings, pdf_bookings, used_idx, fallback_path)
        out_path = fallback_path
        out_name = fallback_name

    # summary text
    from collections import Counter
    status_counts = Counter(eb.match_status for eb in excel_bookings)
    summary_lines = [
        f"Reconciliation complete.",
        f"",
        f"Folder:   {folder}",
        f"Output:   {out_name}",
        f"",
        f"Landlord rows: {len(excel_bookings)}",
        f"  Matched cleanly:        {status_counts.get('MATCH', 0)}",
        f"  Zero-charge (review):   {status_counts.get('ZERO-CHARGE', 0)}",
        f"  Unclear (review):       {status_counts.get('UNCLEAR', 0)}",
        f"  Amount mismatch:        {status_counts.get('MISMATCH', 0)}",
        f"  Missing from PDFs:      {status_counts.get('MISSING', 0)}",
        f"",
        f"PDF bookings not billed:  {sum(1 for p in pdf_bookings if not p.billed_on_bill)}",
    ]
    summary = "\n".join(summary_lines)
    print()
    print(summary)
    return summary, out_path


def main():
    print("=" * 60)
    print("  Aircon Rental Cross-Check")
    print("=" * 60)
    print()

    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = pick_folder()

    if not folder:
        print("No folder selected. Exiting.")
        return

    try:
        summary, out_path = reconcile(folder)
        show_message("Reconciliation Complete",
                     summary + f"\n\nOpen the output file:\n{out_path}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        show_message("Reconciliation Failed", str(e), is_error=True)


if __name__ == "__main__":
    main()
